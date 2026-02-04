"""
    Copyright (C) 2025-26 Dipl.-Ing. Christoph Massmann <chris@dev-investor.de>

    This file is part of pp-terminal.

    pp-terminal is free software: you can redistribute it and/or modify
    it under the terms of the GNU General Public License as published by
    the Free Software Foundation, either version 3 of the License, or
    (at your option) any later version.

    pp-terminal is distributed in the hope that it will be useful,
    but WITHOUT ANY WARRANTY; without even the implied warranty of
    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
    GNU General Public License for more details.

    You should have received a copy of the GNU General Public License
    along with pp-terminal. If not, see <http://www.gnu.org/licenses/>.
"""

from pathlib import Path
from typing import Any
import uuid

import pandas as pd
from openpyxl.styles import Font, PatternFill, numbers
from openpyxl.utils import get_column_letter

from .strategy import OutputStrategy
from .table_decorator import TableOptions, camel_case_to_title
from ..domain.schemas import Money

_HEADER_BG_COLOR = 'D3D3D3'

class ExcelOutputStrategy(OutputStrategy):
    def __init__(self) -> None:
        self._filename: Path | None = None
        self._sheet_counter = 0

    def result_table(self, df: pd.DataFrame | None, options: TableOptions) -> Any:  # pylint: disable=too-many-locals
        if df is None:
            return self.empty_result()

        # Generate filename on first call
        if self._filename is None:
            unique_id = uuid.uuid4().hex[:8]
            self._filename = Path.cwd() / f"output_{unique_id}.xlsx"

        self._sheet_counter += 1
        sheet_name = self._sanitize_sheet_name(options.title) if options.title else f"Sheet{self._sheet_counter}"

        formatted_df = self._prepare_dataframe(df, options)

        mode = 'a' if self._filename.exists() else 'w'
        if_sheet_exists = 'replace' if mode == 'a' else None

        with pd.ExcelWriter(
            self._filename,
            engine='openpyxl',
            mode=mode,
            if_sheet_exists=if_sheet_exists
        ) as writer:
            formatted_df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=options.show_index
            )

            worksheet = writer.sheets[sheet_name]
            self._apply_formatting(worksheet, formatted_df, options, df)

        if self._sheet_counter == 1:
            return (f"Created Excel file: {self._filename}\n", )
        return None

    @staticmethod
    def _prepare_dataframe(df: pd.DataFrame, options: TableOptions) -> pd.DataFrame:
        """Prepare DataFrame by adding totals row if needed."""
        formatted_df = df.copy()

        if options.show_total:
            summary_row = df.select_dtypes(include='number').sum()
            if 'currency' in df.columns:
                summary_row['currency'] = df['currency'].mode()[0] if not df['currency'].mode().empty else ''

            # Add 'Total' label to first non-currency column
            first_col = next((col for col in df.columns if col != 'currency'), None)
            if first_col:
                summary_row[first_col] = 'Total'

            formatted_df = pd.concat([formatted_df, summary_row.to_frame().T], ignore_index=True)

        if 'currency' in formatted_df.columns:
            formatted_df = formatted_df.drop(columns=['currency'])

        return formatted_df

    @staticmethod
    def _apply_formatting(worksheet: Any, formatted_df: pd.DataFrame, options: TableOptions, original_df: pd.DataFrame) -> None:  # pylint: disable=too-many-locals
        """Apply rich formatting to the worksheet."""
        # Format header row
        gray_fill = PatternFill(start_color=_HEADER_BG_COLOR, end_color=_HEADER_BG_COLOR, fill_type='solid')
        for idx, cell in enumerate(worksheet[1]):
            cell.font = Font(bold=True)
            cell.fill = gray_fill
            # Apply camel_case_to_title to column headers (skip index column if present)
            if options.show_index and idx == 0:
                continue  # Keep index header as-is
            col_idx = idx - 1 if options.show_index else idx
            if col_idx < len(formatted_df.columns):
                cell.value = camel_case_to_title(str(formatted_df.columns[col_idx]))

        # Freeze top row
        worksheet.freeze_panes = 'A2'

        # Auto-adjust column widths and apply number formats
        for idx, column in enumerate(formatted_df.columns, start=1):
            column_letter = get_column_letter(idx)
            max_length = max(
                len(str(column)),
                formatted_df[column].astype(str).str.len().max()
            )
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

            # Apply currency formatting to Money columns
            if column in original_df.columns and len(original_df) > 0:
                sample_value = original_df[column].iloc[0]
                if isinstance(sample_value, Money):
                    # Apply currency format to data cells (skip header)
                    for row_idx in range(2, len(formatted_df) + 2):
                        cell = worksheet[f"{column_letter}{row_idx}"]
                        cell.number_format = numbers.FORMAT_CURRENCY_EUR_SIMPLE

        # Bold totals row if present
        if options.show_total and len(formatted_df) > 0:
            total_row_idx = len(formatted_df) + 1
            for cell in worksheet[total_row_idx]:
                cell.font = Font(bold=True)

    @staticmethod
    def _sanitize_sheet_name(name: str) -> str:
        """Sanitize sheet name for Excel compatibility."""
        invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
        sanitized = name
        for char in invalid_chars:
            sanitized = sanitized.replace(char, '')
        # Truncate to 31 characters (Excel limit)
        return sanitized[:31]
